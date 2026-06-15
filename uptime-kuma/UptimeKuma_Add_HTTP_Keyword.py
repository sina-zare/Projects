import time

from uptime_kuma_api import UptimeKumaApi, MonitorType
from openpyxl import load_workbook


api = UptimeKumaApi('http://172.17.255.85:3016/')
api.login('admin', 'I4=t8K<xn')



# Load the workbook and select the first sheet
workbook = load_workbook("C:/Temp/migration.xlsx")
sheet = workbook.active  # Alternatively, use sheet = workbook['Sheet1']

list_vps = []
# Read all data in the sheet
for row in sheet.iter_rows(values_only=True):
    list_vps.append(row)

# Close the workbook
workbook.close()

for node in list_vps:
    name = f'VRA-{(node[0].split('-')[1]).capitalize()}'
    url = f"https://{node[1]}"
    desc = node[2]



    monitor_config = {
        "type": MonitorType.KEYWORD,
        "name": f"{name}",
        "url": f"{url}",
        "keyword": "ورود کاربران",
        "interval": 90,
        "maxretries": 3,
        "retryInterval": 30,
        "timeout": 72,
        "resendInterval": 480,
        "maxredirects": 10,
        "accepted_statuscodes": ['200-299'],
        "description": f"{desc}",
        "notificationIDList": [1]  # Enable Telegram Notification
    }

    try:
        result = api.add_monitor(**monitor_config)
        print(result, end=' : ')
        print(name)
        time.sleep(0.5)
    except Exception as err:
        print(err)

api.disconnect()
