import time
import openpyxl
from pyvim import connect
from pyvim.connect import Disconnect
from pyVmomi import vim
import ssl
import warnings
import os


# Credentials
from cryptography.fernet import Fernet
def decryptor(enc_env_var, key_env_var):

    # Load the key
    key = os.environ.get(key_env_var)
    encryption_key = Fernet(key)
    encrypted_password = (os.environ.get(enc_env_var)).encode()
    # Decrypt Data
    decrypted_password = encryption_key.decrypt(encrypted_password.decode())

    #print(f"Decryped Text: {decrypted_password}")
    return decrypted_password.decode()

username = decryptor("enc_sinaz_abramad","key_sinaz_abramad")
password = decryptor("enc_sinaz_pass","key_sinaz_pass")

#####################################
############## Excel ################

#Taking CRM file for Namayande information
crm_reports_path = "E:/IaaC/Portal_Notification/CRM_Reports/"
# Get a list of all files in the folder
crm_files = os.listdir(crm_reports_path)
# Sort the files based on their creation time
crm_sorted_files = sorted(crm_files, key=lambda x: os.path.getctime(os.path.join(crm_reports_path, x)), reverse=True)
# Get the latest file (first element in the sorted list)
latest_crm_report = crm_sorted_files[0]

full_path_to_latest_crm_report = crm_reports_path + latest_crm_report
# Load the workbook
workbook = openpyxl.load_workbook(full_path_to_latest_crm_report)
# Select the active sheet
sheet = workbook.active
# Create an empty list to store the rows
crm_rows = []
# Iterate through the rows and append each row to the list
for row in sheet.iter_rows(values_only=True):
    crm_rows.append(row)

valuable_crm_data = []
# Taking NID, Agent Name, Email, Phone No
for data in crm_rows:
    valuable_crm_data.append([data[1], data[5], data[6], data[8]])

######################################


######################################
############## vCenter ###############

# Ignore the warning
warnings.filterwarnings("ignore", category=DeprecationWarning)
# *** Connecting to ME-VC01.Abramad.Com to get the Report ***

# Create an SSL context with no certificate verification
context = ssl.SSLContext(ssl.PROTOCOL_TLS)
context.verify_mode = ssl.CERT_NONE

vcenter = connect.SmartConnect(host='me-vc01.abramad.com', user=username, pwd=password, port=443, sslContext=context)
content = vcenter.RetrieveContent()
vm_view = content.viewManager.CreateContainerView(content.rootFolder, [vim.VirtualMachine], True)
vms = [vm for vm in vm_view.view if (vm.name.startswith("MER-") or vm.name.startswith("MERD-") or vm.name.startswith("MEF-") or vm.name.startswith("MES-") or vm.name.startswith("MEA-") or vm.name.startswith("MEB-") or vm.name.startswith("MEM-") or vm.name.startswith("MEI-") or vm.name.startswith("MESA-") or vm.name.startswith("MEV-"))]


customer_vm = input("VM Name: ")
vm_list = []
for vm in vms:

    if vm.name.lower() == customer_vm.lower().strip():

        # Get National ID Status
        vm_national_id = ""
        custom_value_n = vm.summary.customValue
        for i in custom_value_n:
            if i.key == 611:
                vm_national_id = i.value

        vm_list.append([vm.name, vm_national_id])

Disconnect(vcenter)

######################################


######################################
########## Merging 2 lists  ##########

vm_full_info_list = []
for vcenter_item in vm_list:
    for crm_item in valuable_crm_data:
        # if National IDs matched:
        if vcenter_item[1] == crm_item[0]:
            vm_full_info_list.append([vcenter_item[0], vcenter_item[1], crm_item[1], crm_item[2], crm_item[3]])

######################################


######################################
######### Saving Excel File  #########
# Load the Excel file
#workbook_final = openpyxl.Workbook()

# Select the worksheet where you want to add the final data
#worksheet_final = workbook_final.active

# Write first row
#header = ["groupname", "mobile", "address", "Group Value"]
#worksheet_final.append(header)

# Write Final Data
#for final_data in vm_full_info_list:
#    final_row = ["CustomerNationalId", final_data[4], final_data[3], final_data[1] ]
#    worksheet_final.append(final_row)

# Save the changes to the Excel file
#workbook_final.save(f"C:/Users/sina.z/Desktop/Automation_Reports/Portal_Notification/{customer_vm.lower().strip()}.xlsx")

from openpyxl.styles import NamedStyle

# Load the Excel file
workbook_final = openpyxl.Workbook()

# Select the worksheet where you want to add the final data
worksheet_final = workbook_final.active

# Create a text style for the cells
text_style = NamedStyle(name="text_style")
text_style.number_format = '@'  # Text format

# Write first row
header = ["groupname", "mobile", "address", "Group Value"]
worksheet_final.append(header)

# Apply text style to the last column header
worksheet_final.cell(row=1, column=len(header)).style = text_style

# Write Final Data
for idx, final_data in enumerate(vm_full_info_list, start=2):
    final_row = ["CustomerNationalId", final_data[4], final_data[3], final_data[1]]
    worksheet_final.append(final_row)
    # Apply text style to the last column of each row
    worksheet_final.cell(row=idx, column=len(header)).style = text_style

# Save the changes to the Excel file
workbook_final.save(f"E:/IaaC/Portal_Notification/Portal_Template/{customer_vm.lower().strip()}.xlsx")


smiley = '''
                          oooo$$$$$$$$$$$$oooo
                      oo$$$$$$$$$$$$$$$$$$$$$$$$o
                   oo$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$o         o$   $$ o$
   o $ oo        o$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$o       $$ $$ $$o$
oo $ $ "$      o$$$$$$$$$    $$$$$$$$$$$$$    $$$$$$$$$o       $$$o$$o$
"$$$$$$o$     o$$$$$$$$$      $$$$$$$$$$$      $$$$$$$$$$o    $$$$$$$$
  $$$$$$$    $$$$$$$$$$$      $$$$$$$$$$$      $$$$$$$$$$$$$$$$$$$$$$$
  $$$$$$$$$$$$$$$$$$$$$$$    $$$$$$$$$$$$$    $$$$$$$$$$$$$$  """$$$
   "$$$""""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$     "$$$
    $$$   o$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$     "$$$o
   o$$"   $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$       $$$o
   $$$    $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$" "$$$$$$ooooo$$$$o
  o$$$oooo$$$$$  $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$   o$$$$$$$$$$$$$$$$$
  $$$$$$$$"$$$$   $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$     $$$$""""""""
 """"       $$$$    "$$$$$$$$$$$$$$$$$$$$$$$$$$$$"      o$$$
            "$$$o     """$$$$$$$$$$$$$$$$$$"$$"         $$$
              $$$o          "$$""$$$$$$""""           o$$$
               $$$$o                 oo             o$$$"
                "$$$$o      o$$$$$$o"$$$$o        o$$$$
                  "$$$$$oo     ""$$$$o$$$$$o   o$$$$""
                     ""$$$$$oooo  "$$$o$$$$$$$$$"""
                        ""$$$$$$$oo $$$$$$$$$$
                                """"$$$$$$$$$$$
                                    $$$$$$$$$$$$
                                     $$$$$$$$$$"
                                      "$$$""""
'''
print('\n\n' + smiley)
time.sleep(1)