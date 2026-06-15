from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pyvim.connect import Disconnect
from pyvim import connect
from pyVmomi import vim
import openpyxl
import warnings
import smtplib
import time
import csv
import ssl
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
from prometheus_client import CollectorRegistry, Gauge, push_to_gateway, Counter
import traceback
import random
import time
import os
import pyzipper

# --- Configuration ---
script_name = 'downtime_portal_notification'
total_exec_counter_file = f'C://Temp//Script_Metrics//{script_name}-total-execs.txt'
total_failed_exec_counter_file = f'C://Temp//Script_Metrics//{script_name}-total-failed-execs.txt'
pushgateway_url = 'https://me-prometheus.abramad.com:9091'
job_name = 'python_scripts'
instance = script_name
datacenter = 'miremad'
target = 'me_onprem_customers_ticketing'


# Create a registry for our custom metrics
registry = CollectorRegistry()

# Define metrics
duration_gauge = Gauge('script_exec_duration_seconds', 'Duration of my script', registry=registry)
status_gauge = Gauge('script_success', 'Whether script succeeded (1) or failed (0)', registry=registry)
total_execution_counter = Counter('script_total_execs', 'Total number of times the script has run', registry=registry)
total_failed_execution_counter = Counter('script_total_failed_execs', 'Total number of times the script has failed to finish gracefully', registry=registry)
last_error_message = Gauge('script_last_error_message','The last error message encountered during script execution',['error_summary', 'error_detail'], registry=registry)


# Simulate your script logic
start_time = time.time()
success = True
error_string_summary = ""
error_string_detail = ""


try:
    # Zipper function
    def make_zip(files, zip_name, password):
        with pyzipper.AESZipFile(zip_name, 'w', compression=pyzipper.ZIP_LZMA, encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(password.encode())

            for file_path in files:
                file_name = os.path.basename(file_path)  # <- keeps only filename
                zf.write(file_path, arcname=file_name)  # <- overrides path inside ZIP


    # --- Read script run counter from file ---
    def read_value_from_file(file_path):
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)  # Create the directory if it doesn't exist

        if not os.path.exists(file_path):
            with open(file_path, 'w') as f:
                f.write('0')
            return 0

        try:
            with open(file_path, 'r') as f:
                return int(f.read().strip())
        except ValueError:
            # In case of a corrupt or non-integer value
            return 0

    # --- Write updated count to file ---
    def write_value_to_file(file_path, value):
        with open(file_path, 'w') as f:
            f.write(str(value))



    def send_anonymous_email(from_email, to_email, cc_email, subject, html_message, direction,
                             mail_server='mail.abramad.com', attachments=None):
        try:
            ##############################################
            ######### HTML Body Begin For Email ##########
            html_line_break = '''
                                    <p><br></p>
                                '''
            html_msg_1 = f'''
                                <html dir={direction}>
                                  <body>
                                '''
            html_msg_2 = f'''
                                    <p  style="font-family: DiodrumArabic-Regular">{html_message}</p>
                                '''
            html_msg_3 = f'''
                                    '''
            html_msg_4 = '''
                                  </body>
                                </html>
                                '''
            ######### HTML Body End For Email ##########
            ############################################

            email_body = html_msg_1 + html_msg_2 + html_line_break + html_line_break + html_msg_3 + html_msg_4

            # Split email addresses into lists
            to_email_list = to_email.split(",") if to_email else []
            cc_email_list = cc_email.split(",") if cc_email else []
            all_recipients = to_email_list + cc_email_list

            # Create the email message
            msg = MIMEMultipart()
            msg["From"] = Header(from_email, "utf-8")
            msg["To"] = Header(", ".join(to_email_list), "utf-8")  # For display purposes
            msg["CC"] = Header(", ".join(cc_email_list), "utf-8")  # For display purposes
            msg["Subject"] = Header(subject, "utf-8")

            # Attach HTML body
            msg.attach(MIMEText(email_body, "html", "utf-8"))

            # Attach files if any
            if attachments:
                for attachment in attachments:
                    if os.path.exists(attachment):
                        with open(attachment, 'rb') as f:
                            part = MIMEApplication(f.read(), Name=os.path.basename(attachment))
                            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment)}"'
                            msg.attach(part)
                    else:
                        print(f"Warning: Attachment {attachment} not found.")

            # Connect to the mail server and send the email
            with smtplib.SMTP(mail_server, 25) as server:
                server.sendmail(from_email, all_recipients, msg.as_string())
                print("Email sent successfully.")

        except Exception as err:
            print(f"email_sender Function Error: {err}")
            send_anonymous_email('ScriptErrors@abramad.com', 'abramadsysops@abramad.com', 'sina.z@abramad.com',
                                 f"email_sender Function Error in running All_VMs_Info.py",
                                 f"Error Occurred:<br><b>{err}<br></b> Agent: Downtime_Portal_Notification.py",
                                 'ltr')


    # Credentials
    from cryptography.fernet import Fernet


    def decryptor(enc_env_var, key_env_var):
        # Load the key
        key = os.environ.get(key_env_var)
        encryption_key = Fernet(key)
        encrypted_password = (os.environ.get(enc_env_var)).encode()
        # Decrypt Data
        decrypted_password = encryption_key.decrypt(encrypted_password.decode())

        # print(f"Decryped Text: {decrypted_password}")
        return decrypted_password.decode()


    username = 'sysops-svc@abramad.com'
    password = decryptor('sysops-svc_enc', 'sysops-svc_key')

    # Zip file path
    zip_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Update_Files.zip"
    zip_pass = "wzo}e!H32u}8}Dln'[k;"

    today = datetime.now().date()
    day_name = today.strftime('%a')
    if ((day_name.lower() == "sat") or (day_name.lower() == "sun") or (day_name.lower() == "mon") or (
            day_name.lower() == "tue") or (day_name.lower() == "wed")):

        # Read the existing version check content from the file
        version_check_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime/Version_Check.txt"
        version_check = ""
        with open(version_check_path, "r") as file:
            version_check = file.read()

        # Read the existing last ticket No created from the file
        version_no_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime/Version.txt"
        with open(version_no_path, "r") as file:
            version_no = file.read()

        if int(version_no) != int(version_check):

            #############################################################
            ########################## CSV ##############################

            # Taking VM names from 'Planned_For_Update_System.csv' to save as a list #

            # Path to csv file
            csv_file = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime/Planned_For_Update_System.csv"


            # Open the CSV file
            with open(csv_file, newline='') as csvfile:
                # Create a CSV reader
                reader = csv.reader(csvfile)
                # Initialize an empty list to store the rows
                to_update_vm_list = []
                # Loop through each row in the CSV file
                for row in reader:
                    # Append the row to the list
                    # Apply lower() to each element in the row and append the modified row to the list
                    to_update_vm_list.append([element.lower() for element in row])

            # Use a nested list comprehension to flatten the list and convert all elements to strings
            to_update_flatten_vm_list = [str(item) for sublist in to_update_vm_list for item in sublist]

            ##############################################################

            ##############################################################
            ########################### Excel ############################

            # Taking CRM file for Namayande information
            crm_reports_path = "C:/Users/sina.z/Desktop/CRM Reports/"
            # Get a list of all files in the folder
            crm_files = os.listdir(crm_reports_path)
            # Sort the files based on their creation time
            crm_sorted_files = sorted(crm_files, key=lambda x: os.path.getctime(os.path.join(crm_reports_path, x)),
                                      reverse=True)
            # Get the latest file (first element in the sorted list)
            latest_crm_report = crm_sorted_files[0]

            full_path_to_latest_crm_report = crm_reports_path + latest_crm_report
            # Load the workbook
            workbook = openpyxl.load_workbook(full_path_to_latest_crm_report)
            # Select the active sheet
            sheet = workbook.active
            # Create an empty list to store the rows
            crm_rows = []
            # Iterate through the rows and append each row to the list
            for row in sheet.iter_rows(values_only=True):
                crm_rows.append(row)

            valuable_crm_data = []
            # Taking NID, Agent Name, Email, Phone No
            for data in crm_rows:
                valuable_crm_data.append([data[1], data[5], data[6], data[8]])

            ##############################################################

            ##############################################################
            ########################### vCenter ##########################

            # Ignore the warning
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            # *** Connecting to ME-VC01.Abramad.Com to get the Report ***

            # Create an SSL context with no certificate verification
            context = ssl.SSLContext(ssl.PROTOCOL_TLS)
            context.verify_mode = ssl.CERT_NONE

            vcenter = connect.SmartConnect(host='me-vc01.abramad.com', user=username, pwd=password, port=443,
                                           sslContext=context)
            content = vcenter.RetrieveContent()
            vm_view = content.viewManager.CreateContainerView(content.rootFolder, [vim.VirtualMachine], True)
            vcenter_vms = [vm for vm in vm_view.view if (
                        vm.name.startswith("MER-") or vm.name.startswith("MERD-") or vm.name.startswith(
                    "MEF-") or vm.name.startswith("MES-") or vm.name.startswith("MEA-") or vm.name.startswith(
                    "MEB-") or vm.name.startswith("MEM-") or vm.name.startswith("MEI-") or vm.name.startswith(
                    "MESA-") or vm.name.startswith("MEV-"))]

            to_update_vm_list_detailed = []

            for vm_in_update_file in to_update_flatten_vm_list:
                for vm_in_vcenter in vcenter_vms:
                    if vm_in_update_file == vm_in_vcenter.name.lower():

                        # Get National ID Status
                        vm_national_id = ""
                        custom_value_n = vm_in_vcenter.summary.customValue
                        for i in custom_value_n:
                            if i.key == 611:
                                vm_national_id = i.value

                        to_update_vm_list_detailed.append([vm_in_vcenter.name, vm_national_id])

            Disconnect(vcenter)

            ##############################################################

            ##############################################################
            ##################### Merging 2 lists  #######################

            vm_full_info_list = []
            for vcenter_item in to_update_vm_list_detailed:
                for crm_item in valuable_crm_data:
                    # if National IDs matched:
                    if vcenter_item[1] == crm_item[0]:
                        vm_full_info_list.append([vcenter_item[0], vcenter_item[1], crm_item[1], crm_item[2], crm_item[3]])

            # Convert inner lists to tuples
            vm_full_info_tuples = [tuple(inner_list) for inner_list in vm_full_info_list]

            # Create a set of tuples to remove duplicates
            unique_tuples = set(vm_full_info_tuples)

            # Convert tuples back to lists
            vm_full_info_list = [list(inner_tuple) for inner_tuple in unique_tuples]

            # Use a dictionary to keep track of unique items by the first element
            unique_data = {}
            for item in vm_full_info_list:
                unique_data[item[1]] = item

            # Convert the dictionary values back to a list
            unique_list = list(unique_data.values())

            ##############################################################

            ##############################################################
            #################### Saving Excel File #######################

            from openpyxl.styles import NamedStyle

            # Load the Excel file
            workbook_final = openpyxl.Workbook()

            # Select the worksheet where you want to add the final data
            worksheet_final = workbook_final.active

            # Create a text style for the cells
            text_style = NamedStyle(name="text_style")
            text_style.number_format = '@'  # Text format

            # Write first row
            header = ["groupname", "mobile", "address", "Group Value"]
            worksheet_final.append(header)

            # Apply text style to the last column header
            worksheet_final.cell(row=1, column=len(header)).style = text_style

            # Write Final Data
            for idx, final_data in enumerate(vm_full_info_list, start=2):
                final_row = ["CustomerNationalId", final_data[4], final_data[3], final_data[1]]
                worksheet_final.append(final_row)
                # Apply text style to the last column of each row
                worksheet_final.cell(row=idx, column=len(header)).style = text_style

            # Save the changes to the Excel file
            workbook_final.save("C:/Users/sina.z/Desktop/Automation_Reports/Downtime/Portal_Notification.xlsx")
            portal_excel = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime/Portal_Notification.xlsx"
            # Zipping files with password
            files = [portal_excel]
            make_zip(files, zip_path, zip_pass)

            ##############################################################
            ###################### Sending Email #########################

            # Sending Email part

            # Email Recipients
            receiver_email = 'itsm@abramad.com'
            cc_email = 'support@abramad.com,sales@abramad.com,noc@abramad.com'

            html_line_break = '''
                            <p><br></p>
            '''
            html_msg_1s = '''
                            <html dir="rtl">
                            <head>
                                <style>
                                .numeric_class {
                                    direction: ltr;
                                    font-family: Calibri;
                                    text-align: right;
                                }
    
                                </style>
                              </head>
                              <body>
                            '''
            html_msg_2s = '''
                                <p  style="font-family: DiodrumArabic-Regular">با سلام و احترام</p>
                            '''
            html_msg_3s = f'''
                                <p  style="font-family: DiodrumArabic-Regular"> فایل مربوط به اطلاع رسانی سری {int(version_no) - 1} مشترکین، به پیوست قرار گرفت. ممنون میشم با استفاده از این فایل اطلاع رسانی به مشترکین پنل را انجام دهید.  </p>
                            '''
            html_msg_4s = f'''
                                <p  style="font-family: DiodrumArabic-Regular">با سپاس فراوان</p>
                            '''
            html_msg_5s = f'''
                                <p style="font-family: DiodrumArabic-Regular"></p>
                            '''
            html_msg_6s = '''
                              </body>
                            </html>
                            '''

            inform_email_body = html_msg_1s + html_msg_2s + html_msg_3s + html_line_break + html_msg_4s + html_line_break + html_msg_5s + html_msg_6s


            # Send email function with retry logic

            def send_email_with_retry(retries=5, delay=60):
                for attempt in range(retries):
                    try:
                        send_anonymous_email(
                            from_email="Downtime@abramad.com",
                            to_email=receiver_email,
                            cc_email=cc_email,
                            subject=f'فایل اطلاع رسانی مشترکین پنل | سری {int(version_no) - 1} ام بروزرسانی',
                            html_message=inform_email_body,
                            direction="rtl",
                            attachments=[zip_path]
                        )
                        print("Email sent successfully")
                        break
                    except smtplib.SMTPSenderRefused as e:
                        print(f"Error: {e}")
                        if attempt < retries - 1:
                            print(f"Retrying in {delay} seconds...")
                            time.sleep(delay)
                        else:
                            print("Failed to send email after several attempts")


            send_email_with_retry()

            # Saving the Version no in the Version_Check file
            with open(version_check_path, "w") as file:
                file.write(version_no)

except Exception as err:
    print(f"Script failed: {err}")
    success = False
    error_string_summary = f"{type(err).__name__}: {err}"

    # Get the traceback and extract the last traceback frame
    tb = traceback.extract_tb(err.__traceback__)
    last_call = tb[-1]  # the last traceback frame, where the exception occurred
    error_string_detail = f"Error occurred in file {last_call.filename}, line {last_call.lineno}: {last_call.line}"



finally:
    # Finalizing Metrics
    # Script Duration
    duration = time.time() - start_time
    duration_gauge.set(duration)

    #Script Success Status
    status_gauge.set(1 if success else 0)

    # Script Total Executions
    total_exec_counts = read_value_from_file(total_exec_counter_file) + 1
    write_value_to_file(total_exec_counter_file, total_exec_counts)
    total_execution_counter.inc(total_exec_counts)

    if not success:
        # Script Total Failed Executions
        total_failed_exec_counts = read_value_from_file(total_failed_exec_counter_file) + 1
        write_value_to_file(total_failed_exec_counter_file, total_failed_exec_counts)
        total_failed_execution_counter.inc(total_failed_exec_counts)

        # Script Last Error Message
        last_error_message.labels(error_summary=error_string_summary, error_detail=error_string_detail).set(1)

    elif success:
        # Script Total Failed Executions
        total_failed_exec_counts = read_value_from_file(total_failed_exec_counter_file)
        total_failed_execution_counter.inc(total_failed_exec_counts)

        # Script Last Error Message
        last_error_message.labels(error_summary="None", error_detail="None").set(0)


    # Push metrics to Pushgateway
    push_to_gateway(
        gateway=pushgateway_url,
        job=job_name,
        grouping_key={'instance': instance, 'target': target, 'datacenter': datacenter},
        registry=registry
    )

    print('✅ Metrics Sent.')

