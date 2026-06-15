import csv
# from email.mime.application import MIMEApplication
from pyvim import connect
from pyvim.connect import Disconnect
from pyVmomi import vim
import ssl
import warnings
from persiantools.jdatetime import JalaliDate
from jdatetime import datetime, timedelta
# from email.mime.multipart import MIMEMultipart
# from email.mime.text import MIMEText
# from email.mime.base import MIMEBase
# from email import encoders
from cryptography.fernet import Fernet
# from icalendar import Calendar, Event
import smtplib
# import pandas as pd
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import Header
from prometheus_client import CollectorRegistry, Gauge, push_to_gateway, Counter
import traceback
# import random
import openpyxl
from openpyxl.styles import NamedStyle
import time
import os
import pyzipper

# --- Configuration ---
script_name = 'downtime_vnk'
total_exec_counter_file = f'C://Temp//Script_Metrics//{script_name}-total-execs.txt'
total_failed_exec_counter_file = f'C://Temp//Script_Metrics//{script_name}-total-failed-execs.txt'
pushgateway_url = 'http://me-prometheus.abramad.com:9091'
job_name = 'python_scripts'
instance = script_name
datacenter = 'vanak'
target = 'vnk_onprem_customers'

# Create a registry for our custom metrics
registry = CollectorRegistry()

# Define metrics
duration_gauge = Gauge('script_exec_duration_seconds', 'Duration of my script', registry=registry)
status_gauge = Gauge('script_success', 'Whether script succeeded (1) or failed (0)', registry=registry)
total_execution_counter = Counter('script_total_execs', 'Total number of times the script has run', registry=registry)
total_failed_execution_counter = Counter('script_total_failed_execs',
                                         'Total number of times the script has failed to finish gracefully',
                                         registry=registry)
last_error_message = Gauge('script_last_error_message', 'The last error message encountered during script execution',
                           ['error_summary', 'error_detail'], registry=registry)

# Simulate your script logic
start_time = time.time()
success = True
error_string_summary = ""
error_string_detail = ""

try:

    def normalize_mobile(number):
        if not number or number == 'N/A':
            return "N/A"
        number = number.strip()

        # If it already starts with '0' and length is 11, keep it
        if number.startswith("0") and len(number) == 11:
            return number

        # If it starts with '9' and length is 10, add leading zero
        if number.startswith("9") and len(number) == 10:
            return "0" + number

        # If it's something else, return as-is (or handle differently)
        return number

    # Zipper function
    def make_zip(files, zip_name, password):
        with pyzipper.AESZipFile(zip_name, 'w', compression=pyzipper.ZIP_LZMA, encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(password.encode())

            for file_path in files:
                file_name = os.path.basename(file_path)  # <- keeps only filename
                zf.write(file_path, arcname=file_name)  # <- overrides path inside ZIP


    # --- Read script run counter from file ---
    def read_value_from_file(file_path):
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)  # Create the directory if it doesn't exist

        if not os.path.exists(file_path):
            with open(file_path, 'w') as f:
                f.write('0')
            return 0

        try:
            with open(file_path, 'r') as f:
                return int(f.read().strip())
        except ValueError:
            # In case of a corrupt or non-integer value
            return 0


    # --- Write updated count to file ---
    def write_value_to_file(file_path, value):
        with open(file_path, 'w') as f:
            f.write(str(value))


    def send_anonymous_email(from_email, to_email, cc_email, subject, html_message, direction,
                             mail_server='mail.abramad.com', attachments=None):
        try:
            ##############################################
            ######### HTML Body Begin For Email ##########
            html_line_break = '''
                                    <p><br></p>
                                '''
            html_msg_1 = f'''
                                <html dir={direction}>
                                  <body>
                                '''
            html_msg_2 = f'''
                                    <p  style="font-family: DiodrumArabic-Regular">{html_message}</p>
                                '''
            html_msg_3 = f'''
                                    '''
            html_msg_4 = '''
                                  </body>
                                </html>
                                '''
            ######### HTML Body End For Email ##########
            ############################################

            email_body = html_msg_1 + html_msg_2 + html_line_break + html_line_break + html_msg_3 + html_msg_4

            # Split email addresses into lists
            to_email_list = to_email.split(",") if to_email else []
            cc_email_list = cc_email.split(",") if cc_email else []
            all_recipients = to_email_list + cc_email_list

            # Create the email message
            msg = MIMEMultipart()
            msg["From"] = Header(from_email, "utf-8")
            msg["To"] = Header(", ".join(to_email_list), "utf-8")  # For display purposes
            msg["CC"] = Header(", ".join(cc_email_list), "utf-8")  # For display purposes
            msg["Subject"] = Header(subject, "utf-8")

            # Attach HTML body
            msg.attach(MIMEText(email_body, "html", "utf-8"))

            # Attach files if any
            if attachments:
                for attachment in attachments:
                    if os.path.exists(attachment):
                        with open(attachment, 'rb') as f:
                            part = MIMEApplication(f.read(), Name=os.path.basename(attachment))
                            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment)}"'
                            msg.attach(part)
                    else:
                        print(f"Warning: Attachment {attachment} not found.")

            # Connect to the mail server and send the email
            with smtplib.SMTP(mail_server, 25) as server:
                server.sendmail(from_email, all_recipients, msg.as_string())
                print("Email sent successfully.")

        except Exception as err:
            print(f"email_sender Function Error: {err}")


    def decryptor(enc_env_var, key_env_var):
        # Load the key
        key = os.environ.get(key_env_var)
        encryption_key = Fernet(key)
        encrypted_password = (os.environ.get(enc_env_var)).encode()
        # Decrypt Data
        decrypted_password = encryption_key.decrypt(encrypted_password.decode())

        # print(f"Decryped Text: {decrypted_password}")
        return decrypted_password.decode()


    # Function to check if a VM is under vcenter folder
    def is_in_vc_folder(vm, folder_list):
        parent = vm.parent
        allowed = set(folder_list)  # Faster lookups

        while parent:
            if isinstance(parent, vim.Folder):
                if parent.name in allowed:
                    return True
            parent = parent.parent

        return False


    #####################
    ##### Variables #####
    #####################
    month_dict = {
        "01": "فروردین",
        "02": "اردیبهشت",
        "03": "خرداد",
        "04": "تیر",
        "05": "مرداد",
        "06": "شهریور",
        "07": "مهر",
        "08": "آبان",
        "09": "آذر",
        "10": "دی",
        "11": "بهمن",
        "12": "اسفند"
    }

    week_dict = {
        "Sat": "شنبه",
        "Sun": "یکشنبه",
        "Mon": "دوشنبه",
        "Tue": "سه شنبه",
        "Wed": "چهارشنبه",
        "Thu": "پنجشنبه",
        "Fri": "جمعه"
    }

    key_map = {
        301: "vm_creation_date",
        311: "vm_shutdown_date",
        310: "vm_persian_name",
        302: "vm_public_ip",
        501: "vm_shutdown_ticket_no",
        306: "vm_national_no",
        101: "vm_tafsil_no",
        102: "vm_backup_status",
        303: "vm_url",
        304: "vm_ipsids_status",
        305: "vm_in_dept_status",
        307: "vm_not_monitored_status",
        308: "vm_owner",
        309: "vm_dongle_status",
        312: "vm_site_to_site_status",
        313: "vm_usage",
        314: "vm_vip_status",
        315: "vm_waf_status",
        316: "vm_zabbix_vip_status",
        402: "vm_creation_ticket_no",
        620: "vm_company_name",
        602: "vm_rep_name",
        603: "vm_rep_no",
        604: "vm_rep_mail",
        608: "vm_confidentiality_weight",
        609: "vm_integrity_weight",
        610: "vm_availability_weight",
        611: "vm_product_line",
    }

    update_date = '20'
    from_email_address = "Downtime-VNK@abramad.com"
    update_limit = 120
    prefixes = ("vr1-", "vr2-", "vr3-", "vrd-", "vat-", "vbi-", "vmi-", "vsp-", "vsf-")
    allowed_folders = ["Rahkaran", "BI", "ManagedIaaS", "AutomationEdari", "Sepidar"]
    vcenter_addr = 'vab-vc01.abramad.com'
    username = 'sysops-svc@abramad.com'
    password = decryptor('sysops-svc_enc', 'sysops-svc_key')

    # Read the existing Mediocre flag from the file
    mediocre_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Mediocre.txt"
    mediocre_flag = ""
    with open(mediocre_path, "r") as file:
        mediocre_flag = file.read()

    # Read the existing VIP flag from the file
    vip_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VIP.txt"
    vip_flag = ""
    with open(vip_path, "r") as file:
        vip_flag = file.read()

    # Read the existing last ticket No created from the file
    version_no_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Version.txt"
    with open(version_no_path, "r") as file:
        version_no = file.read()

        # Convert the file contents to an integer

        if version_no.isnumeric():
            int_version_no = int(version_no)
            print("File contents type check good, Value:", int_version_no, "\n\n")

        else:

            send_anonymous_email(
                from_email="ScriptErrors@abramad.com",
                to_email='abramadsysops@abramad.com',
                cc_email='sina.z@abramad.com',
                subject=f'Error in Version Data',
                html_message=f"<p><b>Version File does not contain valid integer data<br><br>Wrong Value: {version_no}</p><br><br>Agent: {script_name}",
                direction="rtl"
                # attachments=[attachment]
            )

    # Get today's Jalali date
    today = datetime.now().date()
    day_name = today.strftime('%a')
    # Add three days to today's date
    downtime_date = today + timedelta(days=3)

    downtime_year = str(downtime_date)[:4]
    downtime_month = str(downtime_date)[5:7]
    downtime_day = str(downtime_date)[8:10]
    downtime_month_label = month_dict[downtime_month]
    downtime_day_label = downtime_date.strftime('%a')
    downtime_day_label_fa = week_dict[downtime_day_label]

    from datetime import datetime, timedelta, date

    # Get today's date
    miladi_today = date.today()
    miladi_current_day = str(miladi_today)[-2:]
    print(f'Current day: {miladi_current_day}')

    # Ignore the warning
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    # *** Connecting to ME-VC01.Abramad.Com to get the Report ***
    # Create an SSL context with no certificate verification
    context = ssl.SSLContext(ssl.PROTOCOL_TLS)
    context.verify_mode = ssl.CERT_NONE

    # Connecting to vCenter
    vc = connect.SmartConnect(host=vcenter_addr, user=username, pwd=password, port=443, sslContext=context)
    content = vc.RetrieveContent()
    vm_view = content.viewManager.CreateContainerView(content.rootFolder, [vim.VirtualMachine], True)
    vms = [vm for vm in vm_view.view if (is_in_vc_folder(vm, allowed_folders) and vm.name.lower().startswith(prefixes))]

    # Sort the me_vms list based on VM names
    sorted_vms = sorted(vms, key=lambda vm: vm.name.lower())

    # Zip file path
    zip_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Update_Files.zip"
    zip_pass = "wzo}e!H32u}8}Dln'[k;"

    # Portal Excel path
    portal_excel_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Portal_Notification.xlsx"

    # CSV File Paths
    vr1_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VR/VR1-Total-VMs.csv"
    vr2_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VR/VR2-Total-VMs.csv"
    vr3_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VR/VR3-Total-VMs.csv"
    vrd_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VR/VRD-Total-VMs.csv"

    vsp_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VSP/VSP-Total-VMs.csv"
    vbi_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VBI/VBI-Total-VMs.csv"
    vat_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VAT/VAT-Total-VMs.csv"
    vmi_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VMI/VMI-Total-VMs.csv"
    vsf_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/VSF/VSF-Total-VMs.csv"

    # mer_vip_ramak_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/MER/MER-VIP-Ramak-Total-VMs.csv"
    # mer_vip_alaziz_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/MER/MER-VIP-ALaziz-Total-VMs.csv"
    # mer_vip_domino_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/MER/MER-VIP-Domino-Total-VMs.csv"
    # mer_vip_tatpsa_total_vms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/MER/MER-VIP-TatPSA-Total-VMs.csv"

    planned_for_update_system_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Planned_For_Update_System.csv"
    planned_for_update_sms_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Planned_For_Update_SMS.csv"
    # raw_sms_data_path = f"C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/raw.xlsx"
    # take out updated vms and store 'Em in a list
    updated_vms_csv_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Updated_VMs.csv"
    updated_vms = []

    # Email Recipients
    #reciever_email = 'support@abramad.com,noc@abramad.com,system@abramad.com,farnaz.sh@abramad.com,SaharP@systemgroup.net,JavadK@systemgroup.net,HesamR@systemgroup.net,NedaPo@systemgroup.net,MozhganFA@systemgroup.net,YahyaM@systemgroup.net,fahimegh@systemgroup.net,ArashP@systemgroup.net,MohammadRam@systemgroup.net,abbasas@systemgroup.net,itsm@abramad.com'
    reciever_email = 'support@abramad.com,system@abramad.com'
    cc_email = 'sina.z@abramad.com'
    finished_email_list = reciever_email

    if miladi_current_day == update_date:
        try:
            # Rahkaran VR1
            with open(vr1_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vr1-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Rahkaran VR1 VMs gathered in {vr1_total_vms_path}')

            # Rahkaran VR2
            with open(vr2_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vr2-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Rahkaran VR2 VMs gathered in {vr2_total_vms_path}')

            # Rahkaran VR3
            with open(vr3_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vr3-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Rahkaran VR3 VMs gathered in {vr3_total_vms_path}')

            # Rahkaran VRD
            with open(vrd_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vrd-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Rahkaran VRD VMs gathered in {vrd_total_vms_path}')

            """
            # Rahkaran VIP
            # Ramak
            with open(mer_vip_ramak_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)

                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])

                for vm in sorted_vms:

                    if vm.name.lower().startswith("mer-ramak") and vm.runtime.powerState.lower() == "poweredon":
                        # Get National ID Status
                        vm_national_id = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 611:
                                vm_national_id = i.value

                        # Get VM Persian Name
                        vm_company_name = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 103:
                                vm_company_name = i.value

                        writer.writerow([vm.name, f"\t{str(vm_national_id)}", vm_company_name])

            # ALaziz
            with open(mer_vip_alaziz_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)

                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])

                for vm in sorted_vms:

                    if vm.name.lower().startswith("mer-alaziz") and vm.runtime.powerState.lower() == "poweredon":
                        # Get National ID Status
                        vm_national_id = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 611:
                                vm_national_id = i.value

                        # Get VM Persian Name
                        vm_company_name = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 103:
                                vm_company_name = i.value

                        writer.writerow([vm.name, f"\t{str(vm_national_id)}", vm_company_name])

            # Domino
            with open(mer_vip_domino_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)

                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])

                for vm in sorted_vms:

                    if vm.name.lower().startswith("mer-domino") and vm.runtime.powerState.lower() == "poweredon":
                        # Get National ID Status
                        vm_national_id = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 611:
                                vm_national_id = i.value

                        # Get VM Persian Name
                        vm_company_name = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 103:
                                vm_company_name = i.value

                        writer.writerow([vm.name, f"\t{str(vm_national_id)}", vm_company_name])

            # TatPSA
            with open(mer_vip_tatpsa_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)

                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])

                for vm in sorted_vms:

                    if vm.name.lower().startswith("mer-tatpsa") and vm.runtime.powerState.lower() == "poweredon":
                        # Get National ID Status
                        vm_national_id = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 611:
                                vm_national_id = i.value

                        # Get VM Persian Name
                        vm_company_name = ""
                        custom_value_n = vm.summary.customValue
                        for i in custom_value_n:
                            if i.key == 103:
                                vm_company_name = i.value

                        writer.writerow([vm.name, f"\t{str(vm_national_id)}", vm_company_name])

            ####### End of MER #######
            ##########################"""
            # Sepidar
            with open(vsp_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vsp-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Sepidar VMs gathered in {vsp_total_vms_path}')

            # BI
            with open(vbi_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vbi-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'BI VMs gathered in {vbi_total_vms_path}')

            # Automation
            with open(vat_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vat-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Automation VMs gathered in {vat_total_vms_path}')

            # ManagedIaaS
            with open(vmi_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vmi-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Managed IaaS VMs gathered in {vmi_total_vms_path}')

            # Saham Fasl
            with open(vsf_total_vms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["VM Name", "National ID", "Persian Name", "Agent name", "Agent Phone", "Agent Mail"])
                for vm in sorted_vms:
                    if vm.name.lower().startswith('vsf-') and vm.runtime.powerState.lower() == "poweredon":
                        vm_attrs = {}
                        for attr in vm.summary.customValue:
                            if attr.key in key_map:
                                vm_attrs[key_map[attr.key]] = attr.value
                        print(f"""
                        {vm.name}
                            vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                            vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                            vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                            vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                            vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
                        """)

                        vm_rep_no_raw = vm_attrs.get('vm_rep_no', 'N/A')
                        vm_rep_no = normalize_mobile(vm_rep_no_raw)
                        writer.writerow([vm.name,
                                         f"\t{str(vm_attrs.get('vm_national_no', 'N/A'))}",
                                         vm_attrs.get('vm_company_name', 'N/A'),
                                         vm_attrs.get('vm_rep_name', 'N/A'),
                                         vm_rep_no,
                                         vm_attrs.get('vm_rep_mail', 'N/A')])

            print(f'Saham Fasl VMs gathered in {vsf_total_vms_path}')

        except Exception as vc_error:
            print(f"""
            Error: {vc_error}
            {vm.name}
                vm_national_no: {str(vm_attrs.get('vm_national_no', 'N/A'))}
                vm_company_name: {vm_attrs.get('vm_company_name', 'N/A')}
                vm_rep_name: {vm_attrs.get('vm_rep_name', 'N/A')}
                vm_rep_no: {vm_attrs.get('vm_rep_no', 'N/A')}
                vm_rep_mail: {vm_attrs.get('vm_rep_mail', 'N/A')}
            """)

            success = False
            error_string_summary = f"{type(vc_error).__name__}: {vc_error}"

            # Get the traceback and extract the last traceback frame
            tb = traceback.extract_tb(vc_error.__traceback__)
            last_call = tb[-1]  # the last traceback frame, where the exception occurred
            error_string_detail = f"Error occurred in file {last_call.filename}, line {last_call.lineno}: {last_call.line}"

        # Assign Version No to 1
        int_version_no = 1
        # Overwrite the file with the new data
        with open(version_no_path, "w") as file:
            file.write(str(int_version_no))

        # Empty the updated_vms csv file
        with open(updated_vms_csv_path, "w") as file:
            file.truncate()

        # Assign Mediocre Flag to 0
        with open(mediocre_path, "w") as file:
            file.write("0")

        # Assign VIP Flag to 0
        with open(vip_path, "w") as file:
            file.write("0")

        # Disconnect vCenter
        Disconnect(vc)

    else:
        print(f"No new VM taking for today: {miladi_current_day}th day of month.\n")

    vms_to_be_updated = []
    data_csv = [vr1_total_vms_path,
                vr2_total_vms_path,
                vr3_total_vms_path,
                vrd_total_vms_path,
                vsp_total_vms_path,
                vbi_total_vms_path,
                vat_total_vms_path,
                vsf_total_vms_path,
                vmi_total_vms_path,
                ]

    # Create a list to store the data
    vr1_data_list = []
    vr2_data_list = []
    vr3_data_list = []
    vrd_data_list = []
    vsp_data_list = []
    vbi_data_list = []
    vat_data_list = []
    vmi_data_list = []
    vsf_data_list = []

    for datum in data_csv:

        if datum == vr1_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vr1_data_list.append(data)

        if datum == vr2_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vr2_data_list.append(data)

        if datum == vr3_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vr3_data_list.append(data)

        if datum == vrd_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vrd_data_list.append(data)

        if datum == vsp_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vsp_data_list.append(data)

        if datum == vbi_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vbi_data_list.append(data)

        if datum == vat_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vat_data_list.append(data)

        if datum == vmi_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vmi_data_list.append(data)

        if datum == vsf_total_vms_path:
            # Read Data from CSV files
            with open(datum, 'r', encoding='utf-8-sig') as read_file:
                reader = csv.reader(read_file)
                # Skip header
                next(reader)

                # Iterate over the remaining rows
                for row in reader:
                    # Create a tuple or data structure to store the row data
                    data = (row[0], row[1].replace(f"\t", ""), row[2], row[3], row[4], row[5])
                    # Add the data to the list
                    vsf_data_list.append(data)

    vm_data = [vr1_data_list, vr2_data_list, vr3_data_list, vrd_data_list, vsp_data_list, vbi_data_list, vat_data_list,
               vmi_data_list, vsf_data_list]

    with open(updated_vms_csv_path, 'r', encoding='utf-8-sig') as read_file:
        reader = csv.reader(read_file)

        # Iterate over the remaining rows
        for row in reader:
            # Create a tuple or data structure to store the row data
            # data = (row[0], row[1].replace(f"\t", ""), row[2], row[3],row[4],row[5])
            # Convert each element of the row list to lowercase and append to updated_vms
            updated_vms.append([item.lower() for item in row])

    to_be_updated_schedule_table = ''
    planned_for_update_vms = []
    counter = 0
    gate_is_locked = False

    # Flatten List
    updated_vms = [item for sublist in updated_vms for item in sublist]

    # Word Files
    ticket_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Ticket.txt"
    with open(ticket_path, "w", encoding="utf-8") as file:
        ticket = "مشترک گرامی ابرآمد\n"
        ticket += f"با توجه به اقدامات مورد نیاز برای بهبود و افزایش سطح کیفیت خدمات، در تاریخ {downtime_day_label_fa} {downtime_day} {downtime_month_label} ماه {downtime_year} بین ساعت 1:00 تا 5:00 بامداد بدلیل بروز رسانی سیستم عامل سرور احتمال اختلال در سرویس‌های شما وجود دارد. لطفا پس از پایان ساعت مذکور سرویس‌های خود را بررسی کنید. در صورتی که سرویس شما با اختلال مواجه شد، تیم پشتیبانی ابرآمد همواره پاسخگوی شما همراهان گرامی خواهد بود."

        file.write(ticket)

    sms_path = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/SMS.txt"
    with open(sms_path, "w", encoding="utf-8") as file:
        sms = "مشترک گرامی ابرآمد\n"
        sms += f"با توجه به اقدامات مورد نیاز برای بهبود و افزایش سطح کیفیت خدمات، در تاریخ {downtime_day_label_fa} {downtime_day} {downtime_month_label} ماه {downtime_year} بین ساعت 1:00 تا 5:00 بامداد بدلیل بروز رسانی سیستم عامل سرور احتمال اختلال در سرویس‌های شما وجود دارد. لطفا پس از پایان ساعت مذکور سرویس‌های خود را بررسی کنید. در صورتی که سرویس شما با اختلال مواجه شد، تیم پشتیبانی ابرآمد همواره پاسخگوی شما همراهان گرامی خواهد بود."
        sms += f"\nلغو 11"
        sms += f"\nدریافت 12"

        file.write(sms)

    # Initializing Tables
    to_be_updated_schedule_table += """
                                    <table>
                                    <tr>
                                        <td><b>VM Name</b></td>
                                        <td><b>National ID</b></td>
                                        <td><b>Persian Name</b></td>
                                    </tr>
                                    """

    if (day_name.lower() == "sat") or (day_name.lower() == "sun") or (day_name.lower() == "mon") or (
            day_name.lower() == "tue") or (day_name.lower() == "wed"):
        for vm_info in vm_data:

            # Get track of Numbers
            if counter >= update_limit:
                break

            for vm_deeper_data in vm_info:
                if vm_deeper_data[0].lower() in updated_vms:
                    continue
                else:
                    # Get track of Numbers
                    if counter >= update_limit:
                        break
                    # Fill the Table for email sending
                    to_be_updated_schedule_table += f"""
                            <tr>
                                <td>{vm_deeper_data[0]}</td>
                                <td>{vm_deeper_data[1]}</td>
                                <td>{vm_deeper_data[2]}</td>
                            </tr>

                        """

                    # Append Planned_for_update_vms (Name, National ID, Rep Name, Rep Phone, Rep Mail)
                    # print(f'{50 * '#'}\n{vm_deeper_data[0]}, {vm_deeper_data[1]}, {vm_deeper_data[3]}, {vm_deeper_data[4]}, {vm_deeper_data[5]}')
                    planned_for_update_vms.append(
                        [vm_deeper_data[0], vm_deeper_data[1], vm_deeper_data[3], vm_deeper_data[4], vm_deeper_data[5]])

                    # Append Updated_VMs csv
                    with open(updated_vms_csv_path, 'a', newline='', encoding='utf-8_sig') as csv_file:
                        writer = csv.writer(csv_file)
                        writer.writerow([vm_deeper_data[0]])

                    counter += 1

        to_be_updated_schedule_table += "</table>"

        # Filling VM Names for System team
        with open(planned_for_update_system_path, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)

            writer.writerow(["VM Names"])
            for vm_data in planned_for_update_vms:
                writer.writerow([vm_data[0]])

        # Planned_For_Update_SMS Creation
        rows = set()
        for vm_data in planned_for_update_vms:
            rows.add((vm_data[2], vm_data[3], vm_data[4]))  # tuple => dedup
        with open(planned_for_update_sms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["نام", "شماره همراه", "پست الکترونیک"])

            for row in rows:
                writer.writerow(row)

        # Create workbook + sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Create and register text style
        text_style = NamedStyle(name="text_style")
        text_style.number_format = '@'
        workbook.add_named_style(text_style)

        # Write header
        header = ["groupname", "mobile", "address", "Group Value"]
        worksheet.append(header)

        # Apply style to last column header
        worksheet.cell(row=1, column=len(header)).style = text_style

        # Write data rows
        for idx, vm_data in enumerate(planned_for_update_vms, start=2):
            excel_row = ["CustomerNationalId", vm_data[3], vm_data[4], vm_data[1]]
            worksheet.append(excel_row)

            # Apply text style to the last column
            worksheet.cell(row=idx, column=len(header)).style = text_style

        # Save the file
        workbook.save(portal_excel_path)


        '''# Filling SMS Data
        raw_sms_data = pd.read_excel(raw_sms_data_path, dtype=str)

        # Specify the column indices you want to extract (0-based index)
        columns_indices = [1, 3, 5, 6, 8]

        # Extract the specified columns from each row and store them in a list
        extracted_data = [list(row.iloc[columns_indices]) for index, row in raw_sms_data.iterrows()]

        with open(planned_for_update_sms_path, 'w', newline='', encoding='utf-8_sig') as csv_file:
            writer = csv.writer(csv_file)

            writer.writerow(["نام", "شماره همراه", "پست الکترونیک"])

            for id in planned_for_update_vms:
                for raw_info in extracted_data:

                    column1 = ""
                    column2 = ""
                    column3 = ""

                    if id[1].strip() == str(raw_info[0]).strip():
                        column1 = raw_info[1]
                        column2 = raw_info[4]
                        column3 = f"{raw_info[3]}({raw_info[2]})"

                        writer.writerow([column1, column2, column3])

        # End of SMS Calculations

        # Filling Ticket HID for Abramad Support

        sarv_data = []

        sarv_csv_file = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/rawt.csv"
        with open(sarv_csv_file, 'r', encoding='utf-8-sig') as read_file:
            reader = csv.reader(read_file)

            # Skip the first 8 lines
            for _ in range(8):
                next(reader)

            for row in reader:
                sarv_data.append([row[0], row[4]])
        '''
        ####################################################

        sms_email_data = []

        sms_email_csv_file = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Planned_For_Update_SMS.csv"
        with open(sms_email_csv_file, 'r', encoding='utf-8-sig') as read_file:
            reader = csv.reader(read_file)

            # Skip the first 1 line
            for _ in range(1):
                next(reader)

            for row in reader:
                temp_mail = row[2].split("(")
                customer_name = row[0]
                sms_email_data.append([temp_mail[0].lower(), customer_name])

        final_data_hid = []
        final_data_email = []

        '''
        for hid in sarv_data:

            for email in sms_email_data:

                if email[0].lower().strip() == hid[0].lower().strip():
                    final_data_hid.append(hid[1])
                    final_data_email.append(hid[0].lower())
                    print(f"Matched: {hid[0].lower()}")
        '''

        email_data = []
        for i in sms_email_data:
            email_data.append(i[0])

        remainder_data = set(email_data) - set(final_data_email)
        remainder_amount = len(set(email_data)) - len(set(final_data_email))

        remainder_customers = []
        for i in remainder_data:
            for j in sms_email_data:
                if i.lower() == j[0].lower():
                    remainder_customers.append(j[1])

        print("\n")
        print(f'Email Addresses: {set(email_data)}')
        print(f'Emails Count: {len(set(email_data))}')
        #print("\n")
        #print(remainder_data)
        #print(set(remainder_customers))
        #print(remainder_amount)
        '''
        hid_csv_file = "C:/Users/sina.z/Desktop/Automation_Reports/Downtime_VNK/Planned_For_Update_HID.csv"

        with open(hid_csv_file, mode='w', newline='', encoding='utf-8_sig') as file:

            # Create a CSV writer object
            writer = csv.writer(file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            for k in set(final_data_hid):
                writer.writerow([k])

            for z in set(remainder_customers):
                writer.writerow([z])

        # End of Ticket HID Calculations
        '''

        # Check Table Emptiness then send Email
        empty_schedule_table = """
                                    <table>
                                    <tr>
                                        <td><b>VM Name</b></td>
                                        <td><b>National ID</b></td>
                                        <td><b>Persian Name</b></td>
                                    </tr>
                                    </table>"""

        if to_be_updated_schedule_table == empty_schedule_table and int(mediocre_flag) == 0:

            # Normal Customers are all updated

            html_line_break = '''
                                <p><br></p>
                            '''
            html_msg_1s = '''
                            <html dir="rtl">
                            <head>
                                <style>
                                .numeric_class {
                                    direction: ltr;
                                    font-family: Calibri;
                                    text-align: right;
                                }

                                </style>
                              </head>
                              <body>
                            '''
            html_msg_2s = '''
                                <p  style="font-family: DiodrumArabic-Regular">با سلام و احترام</p>
                            '''
            html_msg_3s = f'''
                                <p  style="font-family: DiodrumArabic-Regular">به اطلاع میرساندبروزرسانی سیستم عامل و نرم افزار های مشترکین ابرآمد در دیتاسنتر ونک، شامل سرویس های اتوماسیون، راهکاران، سهام فصل، سپیدار، BI و مدیریت شده   راهکاران برای تاریخ {downtime_month_label} {downtime_year} به اتمام رسید.</p>
                            '''
            html_msg_4s = f'''
                                <p  style="font-family: DiodrumArabic-Regular">سپاس از همکاری و همراهی تمام دوستان</p>
                            '''
            html_msg_5s = f'''
                                <p style="font-family: DiodrumArabic-Regular"></p>
                            '''
            html_msg_6s = '''
                              </body>
                            </html>
                            '''

            mediocre_email_body = html_msg_1s + html_msg_2s + html_msg_3s + html_line_break + html_msg_4s + html_line_break + html_msg_5s + html_msg_6s

            send_anonymous_email(
                from_email=from_email_address,
                to_email=finished_email_list,
                cc_email='',
                subject=f'اتمام بروزرسانی OS و نرم افزار های مشترکین دیتاسنتر ونک | {downtime_month_label} {downtime_year}',
                html_message=mediocre_email_body,
                direction="rtl"
                # attachments=[attachment]
            )

            # Bring up mediocre flag
            with open(mediocre_path, "w", encoding="utf-8") as file:

                file.write("1")

            print("All Mediocre customers Updated Successfully")

        elif to_be_updated_schedule_table == empty_schedule_table and int(mediocre_flag) == 1:
            print("All Mediocre are updated and flag is 1")

        else:
            ######### HTML Body Begin For Email ##########
            html_line_break = '''
                        <p><br></p>
                    '''
            html_msg_1s = '''
                    <html dir="rtl">
                    <head>
                        <style>
                        .numeric_class {
                            direction: ltr;
                            font-family: Calibri;
                            text-align: right;
                        }

                        table {
                            font-family: DiodrumArabic-Regular;
                            border-collapse: collapse;
                            margin-left: 0;
                            direction: rtl;
                        }

                        table td {
                            border: 1px solid black;
                            padding: 8px;
                            width: 250px;
                            text-align: right;
                            direction: rtl; 
                        }
                        </style>
                      </head>
                      <body>
                    '''
            html_msg_2s = '''
                        <p  style="font-family: DiodrumArabic-Regular">با سلام و احترام</p>
                    '''
            html_msg_3s = f'''
                        <p  style="font-family: DiodrumArabic-Regular">در راستای بروزرسانی سیستم عامل مشترکین مدیریت شده ی ابرآمد به جهت نصب آخرین آپدیت های مایکروسافت در {downtime_month_label} ماه مشترکین زیر در تاریخ {downtime_day_label_fa} {downtime_day} {downtime_month_label} {downtime_year} از ساعت 01:00 الی 05:00 بامداد آپدیت میگردند.</p>
                    '''
            html_msg_4s = f'''
                        <p  style="font-family: DiodrumArabic-Regular"><b>{to_be_updated_schedule_table}</b></p>
                    '''
            html_msg_5s = f'''
                        <p  style="font-family: DiodrumArabic-Regular">@AbramadSystem لطفا کالکشنی برای مشترکین قرار گرفته شده در فایل Planned_For_Update_System قرار گرفته در پیوست برای بروزرسانی OS و اپلیکیشن ها در تاریخ مذکور ایجاد فرمایید.</p>
                    '''
            html_msg_6s = f'''
                        <p  style="font-family: DiodrumArabic-Regular">@AbramadSupport لطفا اطلاع رسانی به مشترکین به صورت تیکت، با استفاده از متن درون فایل Ticket.txt قرار گرفته به پیوست را انجام دهید.</p>
                    '''
            html_msg_7s = f'''
                        <p  style="font-family: DiodrumArabic-Regular">توجه : لطفا ارسال پیامک و تیکت به مشترکین با استفاده از فایل Portal_Notification.xlsx انجام گردد.</p>
                    '''
            html_msg_8s = f'''
                        <p style="font-family: DiodrumArabic-Regular">توجه: پسورد فایل پیوست در Password Manager با شما به اشتراک گذاشته شده است.</p>
                    '''
            html_msg_9s = '''
                      </body>
                    </html>
                    '''

            ######### HTML Body End For Email ##########
            ############################################

            inform_email_body = html_msg_1s + html_msg_2s + html_msg_3s + html_line_break + html_msg_5s + html_msg_6s + html_msg_7s + html_msg_8s + html_line_break + html_msg_4s + html_line_break + html_msg_9s


            # Zipping files with password
            files = [ticket_path, sms_path, planned_for_update_sms_path, planned_for_update_system_path, portal_excel_path]
            make_zip(files, zip_path, zip_pass)


            send_anonymous_email(
                from_email=from_email_address,
                to_email=reciever_email,
                cc_email=cc_email,
                subject=f'اطلاعیه بروزرسانی مشترکین مدیریت شده ابرآمد زیرساخت ونک | سری {int_version_no} | {downtime_day} {downtime_month_label} {downtime_year}',
                html_message=inform_email_body,
                direction="rtl",
                attachments=[zip_path]
            )

            '''
            ##############################
            # Create and Attach Calendar #
            ##############################


            # Calculate Date for Calendar
            # Get today's Jalali date
            khareji_today = datetime.now().date()

            # Add two days to today's date
            khareji_downtime_date = khareji_today + timedelta(days=3)

            khareji_downtime_day = int(str(khareji_downtime_date)[8:10])
            khareji_downtime_month = int(str(khareji_downtime_date)[5:7])
            khareji_downtime_year = int(str(khareji_downtime_date)[:4])


            # Calendar Text
            text_cal = f"متن مربوط به ایمیل اطلاع رسانی سری {int_version_no}ام\nبرای {downtime_month_label} ماه {downtime_year}  را بررسی فرمایید\nبا سپاس"


            # Meeting details
            subject = f'اطلاعیه بروزرسانی مشترکین مدیریت شده ابرآمد | سری {int_version_no}  | {downtime_day} {downtime_month_label} {downtime_year}'
            start = datetime(khareji_downtime_year, khareji_downtime_month, khareji_downtime_day, 1, 0)  # Meeting start time
            end = start + timedelta(hours=4)  # Meeting end time
            location = 'سرور های مشترکین مربوطه'
            attendees = ['abramadops@systemgroup.net']
            organizer = 'sina.z@abramad.com'

            # Create iCalendar event
            event = Event()
            event.add('summary', subject)
            event.add('dtstart', start)
            event.add('dtend', end)
            event.add('location', location)
            event.add('organizer', organizer)
            event.add('description', text_cal)
            event.add('status', 'CONFIRMED')

            for attendee in attendees:
                event.add('attendee', attendee)

            # Create iCalendar calendar and add the event
            cal = Calendar()
            cal.add_component(event)

            # Generate the iCalendar content as a string
            ical_content = cal.to_ical()

            calendar_part = MIMEBase('text', 'calendar', method='REQUEST')
            calendar_part.set_payload(ical_content)
            encoders.encode_base64(calendar_part)
            calendar_part.add_header('Content-Disposition', 'attachment; filename="meeting.ics"')

            # Attach Calendar file
            msg.attach(calendar_part)


            # Attach Text to email
            msg.attach(MIMEText(inform_email_body, 'html'))

            # Connect to the SMTP server and send the email 465,*587*,25 (mail.systemgroup.net)
            # Send email info
            smtp_server = 'mail.systemgroup.net'
            smtp_port = 587
            smtp_username = username
            smtp_password = password

            # Send email function
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_username, smtp_password)
                server.sendmail('sina.z@abramad.com', reciever_email.split(",") + cc_email.split(','), msg.as_string())
            '''

            # Increment Version No by one
            int_version_no += 1
            # Overwrite the file with the new data
            with open(version_no_path, "w") as file:
                file.write(str(int_version_no))


except Exception as err:
    print(f"Script failed: {err}")
    success = False
    error_string_summary = f"{type(err).__name__}: {err}"

    # Get the traceback and extract the last traceback frame
    tb = traceback.extract_tb(err.__traceback__)
    last_call = tb[-1]  # the last traceback frame, where the exception occurred
    error_string_detail = f"Error occurred in file {last_call.filename}, line {last_call.lineno}: {last_call.line}"

    print(
        f"Type: {type(err).__name__}\nError occurred in file {last_call.filename}, line {last_call.lineno}: {last_call.line}")




finally:
    # Finalizing Metrics
    # Script Duration
    duration = time.time() - start_time
    duration_gauge.set(duration)

    # Script Success Status
    status_gauge.set(1 if success else 0)

    # Script Total Executions
    total_exec_counts = read_value_from_file(total_exec_counter_file) + 1
    write_value_to_file(total_exec_counter_file, total_exec_counts)
    total_execution_counter.inc(total_exec_counts)



    if not success:
        # Script Total Failed Executions
        total_failed_exec_counts = read_value_from_file(total_failed_exec_counter_file) + 1
        write_value_to_file(total_failed_exec_counter_file, total_failed_exec_counts)
        total_failed_execution_counter.inc(total_failed_exec_counts)

        # Script Last Error Message
        last_error_message.labels(error_summary=error_string_summary, error_detail=error_string_detail).set(1)



    elif success:
        # Script Total Failed Executions
        total_failed_exec_counts = read_value_from_file(total_failed_exec_counter_file)
        total_failed_execution_counter.inc(total_failed_exec_counts)

        # Script Last Error Message
        last_error_message.labels(error_summary="None", error_detail="None").set(0)

    # Push metrics to Pushgateway
    push_to_gateway(
        gateway=pushgateway_url,
        job=job_name,
        grouping_key={'instance': instance, 'target': target, 'datacenter': datacenter},
        registry=registry
    )


    print('✅ Metrics Sent.')
