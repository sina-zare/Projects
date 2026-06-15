import openpyxl
from pyvim import connect
from pyvim.connect import Disconnect
from pyVmomi import vim
import ssl
import warnings
import os
from cryptography.fernet import Fernet

def decryptor(enc_env_var, key_env_var):

    # Load the key
    key = os.environ.get(key_env_var)
    encryption_key = Fernet(key)
    encrypted_password = (os.environ.get(enc_env_var)).encode()
    # Decrypt Data
    decrypted_password = encryption_key.decrypt(encrypted_password.decode())

    #print(f"Decryped Text: {decrypted_password}")
    return decrypted_password.decode()



# Taking Needed Data From xlsx
rahkaran_ver_excell_file = 'C:/Users/sina.z/Desktop/Python-Projects/Files_to_work_with/MER_Rahkaran_Servers_Version.xlsx'
# Load the workbook
workbook = openpyxl.load_workbook(rahkaran_ver_excell_file)
# Select the active sheet
sheet = workbook.active
# Create an empty list to store the rows
raw_rahkaran_ver_list = []
# Iterate through the rows and append each row to the list
for row in sheet.iter_rows(values_only=True):
    raw_rahkaran_ver_list.append(row)


rahkaran_ver_list = []
# Taking Local URL, Package Title, Product Version, Release NO
for data in raw_rahkaran_ver_list:
    temp = data[0].split("/")
    rahkaran_ver_list.append([temp[2], data[0], data[1], data[2], data[3]])

#for i in rahkaran_ver_list:
#    print(f"\nFQDN: {i[0]}\nLocal URL: {i[1]}\nPackage Title: {i[2]}\nProduct Version: {i[3]}\nRelease NO: {i[4]}\n")



#####################################################
################ Connecting to vCenter ##############
#####################################################
username = decryptor("enc_sinaz_abramad","key_sinaz_abramad")
password = decryptor("enc_sinaz_pass","key_sinaz_pass")

# *** Connecting to ME-VC01.Abramad.Com to get the Report ***
warnings.filterwarnings("ignore", category=DeprecationWarning)
context = ssl.SSLContext(ssl.PROTOCOL_TLS)
context.verify_mode = ssl.CERT_NONE

# Connecting to vCenter
ME_VC = connect.SmartConnect(host='me-vc01.abramad.com', user=username, pwd=password, port=443, sslContext=context)
me_content = ME_VC.RetrieveContent()
me_vm_view = me_content.viewManager.CreateContainerView(me_content.rootFolder, [vim.VirtualMachine], True)
me_vms = [vm for vm in me_vm_view.view if (vm.name.startswith("MER-") or vm.name.startswith("MERD-"))]
sorted_vms = sorted(me_vms, key=lambda vm: vm.name.lower())

vm_specs = {}

for vm in sorted_vms:
    power_state = vm.runtime.powerState.lower()
    if power_state == "poweredon":

        # retrieve vm Hostname
        vm_hostname = vm.summary.guest.hostName

        # Get VM Persian Name
        vm_persian_name = ""
        custom_value_n = vm.summary.customValue
        for i in custom_value_n:
            if i.key == 103:
                vm_persian_name = i.value

        # Get VM URL
        vm_url = ""
        vm_custom_attr = vm.summary.customValue
        for i in vm_custom_attr:
            if i.key == 604:
                vm_url = i.value

        # Get National ID Status
        vm_national_id = "None"
        custom_value_n = vm.summary.customValue
        for i in custom_value_n:
            if i.key == 611:
                vm_national_id = i.value

        vm_specs[f'{vm_hostname.lower()}'] = [vm.name, vm_url, vm_national_id, vm_persian_name]

Disconnect(ME_VC)


# Calculation

final_dictionary = {}
for vm_spec in vm_specs:
    for data in rahkaran_ver_list:
        if vm_spec == data[0]: # if both FQDNs matched
            final_dictionary[f"{vm_spec}"] = [vm_specs[vm_spec][0], vm_specs[vm_spec][2], data[2], data[3], data[4], data[1], vm_specs[vm_spec][1], vm_specs[vm_spec][3]]

"""
for i in final_dictionary:
    print("\n")
    print("#" * 30)
    print(f'"Server Name": "{final_dictionary[i][0]}"')
    print(f'"National ID": "{final_dictionary[i][1]}"')
    print(f'"Package Title": "{final_dictionary[i][2]}"')
    print(f'"Product Version": "{final_dictionary[i][3]}"')
    print(f'"Release NO": "{final_dictionary[i][4]}"')
    print(f'"Local URL": "{final_dictionary[i][5]}"')
    print(f'"External URL": "{final_dictionary[i][6]}"')
    print(f'"Persian Name": "{final_dictionary[i][7]}"')
    print("#" * 30)
"""

# Load the Excel file
workbook = openpyxl.Workbook()

# Select the worksheet where you want to add data
worksheet = workbook.active

# Write the dictionary data to the worksheet
header = ["Server Name", "National ID", "Package Title", "Product Version", "Release NO", "Local URL", "External URL", "Persian Name"]
worksheet.append(header)

# Append Worksheet
for vm in final_dictionary:
    worksheet.append(final_dictionary[vm])

# Save the changes to the Excel file
workbook.save('C:/Users/sina.z/Desktop/Automation_Reports/Rahkaran_Versions/Rahkaran_Customer_Versions.xlsx')
