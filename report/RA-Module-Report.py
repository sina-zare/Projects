from openpyxl import load_workbook
import openpyxl

# Load the workbook and select the first sheet
workbook = load_workbook('C:/Temp/RA-CloudReports.xlsx')
sheet = workbook.active  # Alternatively, use sheet = workbook['Sheet1']

list_vps = []
# Read all data in the sheet
for row in sheet.iter_rows(values_only=True):
    list_vps.append(row)

# Close the workbook
workbook.close()


# Initialize the final list with the header
final_list = [['VM Name', 'Persian Name', 'URL', 'Pakhsh Module', 'KhordeForushi Module']]
for i in list_vps:

    khorde = 'ندارد'
    if i[24] and 'خرده فروشی' in i[24]:
        khorde = 'دارد'

    pakhsh = 'ندارد'
    if i[24] and 'پخش' in i[24]:
        pakhsh = 'دارد'

    vm_name = 'ندارد'
    if i[2] is not None:
        vm_name = i[2]

    vm_url = ''
    if i[3] is not None:
        vm_url = f'https://{i[3].lower()}'

    vm_persian_name = ''
    if i[1] is not None:
        vm_persian_name = i[1]

    # Add the processed row to the final list
    final_list.append([(vm_name.split('.')[0]), vm_persian_name, vm_url, pakhsh, khorde])



# Load the new list from the text file
with open('C:/Temp/new_list.txt', 'r') as file:
    lines = [line.strip() for line in file]

god = [['VM Name', 'Persian Name', 'URL', 'Pakhsh Module', 'KhordeForushi Module']]
for srv in lines:
    srv = srv.lower().strip()
    for data in final_list:
        vm_name_lower = data[0].lower().strip()
        if srv == vm_name_lower:
            god.append(data)
            break  # No need to continue once a match is found


# Create a new Excel workbook and add the results
workbook = openpyxl.Workbook()
worksheet = workbook.active

for vm in god:
    worksheet.append(vm)

# Save the changes to the Excel file
workbook.save('C:/Temp/RA-Module-Report.xlsx')