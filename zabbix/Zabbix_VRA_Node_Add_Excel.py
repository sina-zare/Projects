try:
    from openpyxl import load_workbook
    from pyzabbix import ZabbixAPI
    from cryptography.fernet import Fernet
    import os
    from pyvim.connect import Disconnect
    from pyvim import connect
    from pyVmomi import vim
    import traceback
    import warnings
    import time
    import ssl
    import traceback
    import os

except Exception as m_err:
    print(f'Module Import Error: {m_err}')

try:
    def decryptor(enc_env_var, key_env_var):

        # Load the key
        key = os.environ.get(key_env_var)
        encryption_key = Fernet(key)
        encrypted_password = (os.environ.get(enc_env_var)).encode()
        # Decrypt Data
        decrypted_password = encryption_key.decrypt(encrypted_password.decode())

        # print(f"Decryped Text: {decrypted_password}")
        return decrypted_password.decode()

except Exception as f_err:
    print(f'Function Error: {f_err}')


# Nodes Information Gathering
password = decryptor("enc_sinaz_pass", "key_sinaz_pass")



# Load the workbook and select the first sheet
workbook = load_workbook("C:/Temp/migration.xlsx")
sheet = workbook.active  # Alternatively, use sheet = workbook['Sheet1']

list_vps = []
# Read all data in the sheet
for row in sheet.iter_rows(values_only=True):
    list_vps.append(row)

# Close the workbook
workbook.close()







# Zabbix server credentials
zabbix_url = "http://zab.abramad.com/zabbix"
zabbix_user = "sina.z"
zabbix_password = password

# Template and Host Group
template_name = "Rahkaran-Abri-Tpl"  # Template name
host_group_name1 = "Rahkaran-Abri-Grp"  # Host group name
host_group_name2 = "ServiceDesk"  # Host group name

# Connect to Zabbix API
zapi = ZabbixAPI(zabbix_url)
zapi.login(zabbix_user, zabbix_password)

# Step 1: Get the template ID
template = zapi.template.get(filter={"host": template_name})
if not template:
    raise Exception(f"Template '{template_name}' not found.")
template_id = template[0]["templateid"]

# Step 2: Get the host group ID
host_group1 = zapi.hostgroup.get(filter={"name": host_group_name1})
if not host_group1:
    raise Exception(f"Host group '{host_group_name1}' not found.")
host_group_id1 = host_group1[0]["groupid"]

host_group2 = zapi.hostgroup.get(filter={"name": host_group_name2})
if not host_group2:
    raise Exception(f"Host group '{host_group_name2}' not found.")
host_group_id2 = host_group2[0]["groupid"]


for node in list_vps[:]:

    try:
        # Host information
        host_name = f'vra-{(node[0].split('-')[1]).lower()}'         # {HOST.HOST}
        host_visible_name = host_name.split("-")[1]                  # {HOST.NAME}
        host_desc = node[2]                                          # Host Description
        host_ext_url = f"https://{node[1].lower()}"                  # Host Ext URL
        host_int_url = f"http://{host_name}.cloud.local"             # Host Int URL

        tags = [
            {"tag": "Note", "value": ""},
            {"tag": "Owner", "value": "ServiceDesk"},
            {"tag": "Ext URL", "value": f"{host_ext_url}"},
            {"tag": "Int URL", "value": f"{host_int_url}"}
        ]

        # Step 3: Create the new host with description and visible name
        new_pon_host = zapi.host.create({
            "host": host_name,                                                         # Technical name of the host (internal name)
            "name": host_visible_name,                                                 # Visible name of the host in the frontend
            "description": host_desc,                                                  # Host description
            "groups": [{"groupid": host_group_id1}, {"groupid": host_group_id2}],      # Host group
            "templates": [{"templateid": template_id}],                                # Template
            "tags": tags
        })

        host_id = new_pon_host['hostids'][0]
        print(f"Host '{host_name}' created with ID {host_id}")

    except Exception as node_error:
        print(f"Error adding node '{host_name}': {node_error}")
        traceback.print_exc()
        continue



